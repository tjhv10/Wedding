import openpyxl
import quopri
from googletrans import Translator

def translate_names(input_file):
    translated_lines = []
    translator = Translator()

    with open(input_file, 'r', encoding='utf-8') as infile:
        for line in infile:
            if line.startswith('N:') or line.startswith('FN:') or line.startswith('NCHARSET:') or line.startswith('FNCHARSET:'):
                parts = line.split(':', 1)
                if len(parts) == 2:
                    name = parts[1].strip()
                    # Check if the name contains any non-English characters
                    if any(ord(c) > 127 for c in name):
                        # If name is not in English, strip non-English or Hebrew characters
                        name = ''.join(char for char in name if char.isalpha() or char in [' ', '☆', 'ⓝ'])
                    else:
                        # If name is in English, translate to Hebrew
                        print(name)
                        name = translator.translate(name, src='en', dest='he').text
                        print(name)
                        print("------------")
                    line = f"{parts[0]}:{name}\n"
                    # print(line)
            translated_lines.append(line)

    # Write translated lines to the output file
    with open("output.vcf", 'w', encoding='utf-8') as outfile:
        outfile.writelines(translated_lines)

def remove_char_from_file(file_path, char_to_remove):
    try:
        with open(file_path, 'r+', encoding="utf-8") as file:
            lines = file.readlines()
            file.seek(0)
            file.truncate()
            for line in lines:
                new_line = line.replace(char_to_remove, '')
                file.write(new_line)
        
        print(f"Character '{char_to_remove}' removed from '{file_path}'.")

    except FileNotFoundError:
        print("File not found.")

def decode_quoted_printable(line):
    x = quopri.decodestring(line.encode('utf-8')).decode('utf-8')
    return x

def decode_and_replace_in_file(file_path):
    updated_lines = []
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            if "=" in line:
                # decoded_line = decode_quoted_printable(line.split(":")[1].strip())
                line = line.replace(line, '')
            updated_lines.append(line)
    
    with open(file_path, 'w', encoding='utf-8') as file:
        file.writelines(updated_lines)
from difflib import SequenceMatcher

def similar(a, b):
    if a is None:
        return 0
    return SequenceMatcher(None, a, b).ratio()*100
    
def read_excel(file_path):
    data = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]  # Get column headers
    name_col = headers.index('name') + 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[name_col - 1]  # Adjust index to 0-based
        data.append(name)
    # Remove None values
    return data

def read_vcf(file_path):
    data = []
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for i in range(0, len(lines)):
            try:
                x= lines[i].split(':')[0].strip()
                if(x=="FNCHARSET" or x=="FN"):
                    name = lines[i].split(':')[-1].strip()
                    phone = lines[i+1].split(':')[-1].strip()
                    data.append((name, phone))
            except IndexError:
                print("Error: Unexpected format in VCF file.")
    print(data.__len__())
    return data


def match_names(names_excel, names_vcf):
    matched_data = {}
    for name_excel in names_excel:
        max_similarity = 0
        matched_phone = None
        for name_vcf, phone_vcf in names_vcf:
            similarity = similar(name_excel, name_vcf)
            if similarity > max_similarity:
                max_similarity = similarity
                matched_phone = phone_vcf
                match_name = name_vcf
        matched_data[name_excel] = (matched_phone, max_similarity,match_name)
    return matched_data

def write_excel(matched_data, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Matched Name", "Phone", "Similarity Percentage"])
    for name, (matched_name, phone, similarity) in matched_data.items():
        sheet.append([name, matched_name, phone, similarity])
    workbook.save(output_file)


# Paths to input and output files
excel_file = 'wedd.xlsx'
vcf_file = 'contactsA.vcf'
output_file = 'output.xlsx'
# translate_names(vcf_file)
names_excel = read_excel(excel_file)
names_vcf = read_vcf(vcf_file)
matched_data = match_names(names_excel, names_vcf)
write_excel(matched_data, output_file)

# print("Matching completed. Output written to:", output_file)